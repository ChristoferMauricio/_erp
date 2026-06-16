# -*- coding: utf-8 -*-
"""
Exportación a Excel del histórico completo.

- Hoja "Datos": una fila por tarea-insumo (Tabla de Excel `tblDatos`), histórico completo.
- Hojas de agregación: espejan las tablas del tablero (Tabla de Excel cada una).
- PivotTables nativas: se inyectan luego a nivel OOXML (ver inject_pivots / pivots.py).

`generate(query_df)` recibe la función query_df(sql) -> pandas.DataFrame del microservicio
(reusa la conexión pg8000 existente) y devuelve los bytes del .xlsx.
"""
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Columnas de la hoja Datos (orden fijo; los pivots referencian estos nombres/índices)
DATA_COLUMNS = [
    "Tarea ID", "Tipo", "Origen", "Subsistema", "Causa Raíz", "Detalle",
    "Nivel", "Zona", "Punto", "Personas", "Tiempo (h)", "HH", "Periodo",
    "Insumo", "Cantidad", "Unidad",
]

# PivotTables nativas a inyectar (espejo de las tablas del tablero, pero re-pivotables)
PIVOTS = [
    {"sheet": "PV Subsistema", "name": "ptSubsistema", "rows": ["Subsistema"], "col": "Tipo",
     "datas": [("Tarea ID", "count", "Tareas")]},
    {"sheet": "PV Tendencia", "name": "ptTendencia", "rows": ["Periodo"], "col": "Tipo",
     "datas": [("Tarea ID", "count", "Tareas")]},
    {"sheet": "PV Causa", "name": "ptCausa", "rows": ["Causa Raíz"], "col": None,
     "datas": [("Tarea ID", "count", "Tareas"), ("HH", "sum", "Suma de HH")]},
    {"sheet": "PV Nivel", "name": "ptNivel", "rows": ["Nivel"], "col": None,
     "datas": [("HH", "sum", "Suma de HH"), ("Tarea ID", "count", "Tareas")]},
    {"sheet": "PV Insumos", "name": "ptInsumos", "rows": ["Insumo"], "col": None,
     "datas": [("Cantidad", "sum", "Suma de Cantidad"), ("HH", "sum", "Suma de HH")]},
    {"sheet": "PV Zonas", "name": "ptZonas", "rows": ["Zona"], "col": None,
     "datas": [("Tarea ID", "count", "Tareas")]},
]

FLAT_SQL = """
SELECT t.id::text                                              AS "Tarea ID",
       ct.nombre                                               AS "Tipo",
       co.nombre                                               AS "Origen",
       COALESCE(cs.codigo, 'DAT')                              AS "Subsistema",
       COALESCE(cr.nombre, 'Mantenimiento Programado')         AS "Causa Raíz",
       t.detalle                                               AS "Detalle",
       COALESCE(u.nivel, 'Interior Mina')                      AS "Nivel",
       COALESCE(u.zona, 'General')                             AS "Zona",
       u.punto                                                 AS "Punto",
       COALESCE(t.cant_personas, 0)                            AS "Personas",
       COALESCE(t.tiempo_horas, 0)::float                      AS "Tiempo (h)",
       (COALESCE(t.cant_personas,0) * COALESCE(t.tiempo_horas,0))::float AS "HH",
       to_char(t.periodo, 'YYYY-MM')                           AS "Periodo",
       i.nombre_normalizado                                    AS "Insumo",
       ti.cantidad::float                                      AS "Cantidad",
       cum.simbolo                                             AS "Unidad"
FROM tarea t
JOIN cat_tipo ct        ON t.tipo_id = ct.id
JOIN cat_origen co      ON t.origen_id = co.id
LEFT JOIN cat_causa_raiz cr ON t.causa_raiz_id = cr.id
LEFT JOIN cat_subsistema cs ON cr.subsistema_id = cs.id
LEFT JOIN ubicacion u   ON t.ubicacion_id = u.id
LEFT JOIN tarea_insumo ti ON ti.tarea_id = t.id
LEFT JOIN insumo i      ON ti.insumo_id = i.id
LEFT JOIN cat_unidad_medida cum ON ti.unidad_medida_id = cum.id
ORDER BY t.id
"""

# --- estilos ---
HDR_FILL = PatternFill("solid", fgColor="0F766E")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _clean(v):
    """NaN/NaT -> None; numpy escalares -> tipos nativos."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):
        try:
            return v.item()
        except Exception:
            return v
    return v


def _write_table(wb, sheet_name, headers, rows, table_name, style="TableStyleMedium2"):
    """Escribe una hoja con encabezado estilizado + Tabla de Excel (ListObject)."""
    ws = wb.create_sheet(title=sheet_name)
    ws.append(list(headers))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border = BORDER
    for r in rows:
        ws.append([_clean(x) for x in r])
    nrows = len(rows)
    ncols = len(headers)
    # anchos de columna razonables
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        maxlen = len(str(headers[c - 1]))
        for r in rows[:200]:
            val = r[c - 1]
            if val is not None:
                maxlen = max(maxlen, len(str(val)))
        ws.column_dimensions[col].width = min(max(maxlen + 2, 10), 48)
    ws.freeze_panes = "A2"
    # Tabla de Excel: requiere al menos una fila de datos
    last = max(nrows + 1, 2)
    ref = "A1:%s%d" % (get_column_letter(ncols), last)
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    return ws


def _aggregations(df):
    """Calcula las tablas que espejan el tablero. Devuelve lista de (hoja, headers, rows, table_name)."""
    out = []
    tareas = df.drop_duplicates(subset=["Tarea ID"]).copy()
    ins = df[df["Insumo"].notna()].copy()

    # Resumen (KPIs)
    total_tareas = len(tareas)
    total_hh = tareas["HH"].fillna(0).sum()
    total_horas = tareas["Tiempo (h)"].fillna(0).sum()
    total_personas = tareas["Personas"].fillna(0).sum()
    total_cant = ins["Cantidad"].fillna(0).sum()
    resumen = [
        ["Total de tareas", total_tareas],
        ["Total Horas-Hombre (HH)", round(float(total_hh), 1)],
        ["Total horas (Tiempo)", round(float(total_horas), 1)],
        ["Promedio personas/tarea", round(float(total_personas / total_tareas), 2) if total_tareas else 0],
        ["Promedio horas/tarea", round(float(total_horas / total_tareas), 2) if total_tareas else 0],
        ["Líneas de insumo", len(ins)],
        ["Cantidad total de insumos", round(float(total_cant), 2)],
    ]
    out.append(("Resumen", ["Indicador", "Valor"], resumen, "tblResumen"))

    # Tareas x Subsistema
    g = tareas.groupby("Subsistema").size().sort_values(ascending=False)
    out.append(("Tareas x Subsistema", ["Subsistema", "Tareas"],
                [[k, int(v)] for k, v in g.items()], "tblSubsistema"))

    # Tendencia mensual (Periodo x Tipo)
    tt = tareas.dropna(subset=["Periodo"])
    piv = tt.pivot_table(index="Periodo", columns="Tipo", values="Tarea ID",
                         aggfunc="count", fill_value=0).sort_index()
    rows = []
    for periodo, row in piv.iterrows():
        inc = int(row.get("Incidente", 0)); req = int(row.get("Requerimiento", 0))
        rows.append([periodo, inc, req, inc + req])
    out.append(("Tendencia Mensual", ["Periodo", "Incidentes", "Requerimientos", "Total"],
                rows, "tblTendencia"))

    # HH x Causa Raíz
    g = tareas.groupby("Causa Raíz").agg(
        tareas=("Tarea ID", "count"), hh=("HH", "sum"), horas=("Tiempo (h)", "sum")
    ).sort_values("hh", ascending=False)
    out.append(("HH x Causa", ["Causa Raíz", "Tareas", "HH", "Horas"],
                [[k, int(r.tareas), round(float(r.hh), 1), round(float(r.horas), 1)] for k, r in g.iterrows()],
                "tblHHCausa"))

    # HH x Nivel
    g = tareas.groupby("Nivel").agg(
        hh=("HH", "sum"), tareas=("Tarea ID", "count"), horas=("Tiempo (h)", "sum")
    ).sort_values("hh", ascending=False)
    out.append(("HH x Nivel", ["Nivel", "HH", "Tareas", "Horas"],
                [[k, round(float(r.hh), 1), int(r.tareas), round(float(r.horas), 1)] for k, r in g.iterrows()],
                "tblHHNivel"))

    # Causa -> Detalle
    cd = tareas.dropna(subset=["Detalle"]).groupby(["Causa Raíz", "Detalle"]).size().sort_values(ascending=False)
    out.append(("Causa x Detalle", ["Causa Raíz", "Detalle", "Tareas"],
                [[k[0], k[1], int(v)] for k, v in cd.items()], "tblCausaDetalle"))

    # Zonas calientes
    z = tareas.groupby(["Nivel", "Zona"]).size().sort_values(ascending=False)
    out.append(("Zonas Calientes", ["Nivel", "Zona", "Intervenciones"],
                [[k[0], k[1], int(v)] for k, v in z.items()], "tblZonas"))

    # Consumo x Insumo (cantidad + HH asociado)
    gi = ins.groupby("Insumo").agg(cant=("Cantidad", "sum"), hh=("HH", "sum")).sort_values("cant", ascending=False)
    out.append(("Consumo x Insumo", ["Insumo", "Cantidad", "HH asociado"],
                [[k, round(float(r.cant), 2), round(float(r.hh), 1)] for k, r in gi.iterrows()],
                "tblInsumos"))

    # Suministros x Unidad
    su = ins.groupby("Unidad")["Cantidad"].sum().sort_values(ascending=False)
    out.append(("Suministros x Unidad", ["Unidad", "Cantidad"],
                [[k, round(float(v), 2)] for k, v in su.items()], "tblUnidad"))

    return out


def build_base_workbook(df, with_pivots=True):
    """Workbook openpyxl con Datos (Tabla) + hojas de agregación + hojas PV vacías (placeholders para pivots)."""
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # Hoja Datos
    rows = [[_clean(v) for v in rec] for rec in df[DATA_COLUMNS].itertuples(index=False, name=None)]
    _write_table(wb, "Datos", DATA_COLUMNS, rows, "tblDatos", style="TableStyleMedium9")

    # Hojas de agregación
    for sheet, headers, agg_rows, tname in _aggregations(df):
        _write_table(wb, sheet, headers, agg_rows, tname)

    # Hojas PV vacías (placeholders donde se inyectan las PivotTables nativas)
    for cfg in (PIVOTS if with_pivots else []):
        ws = wb.create_sheet(title=cfg["sheet"])
        ws["A1"] = "Tabla dinámica (PivotTable) — se actualiza al abrir en Excel. Si está vacía: clic derecho → Actualizar."
        ws["A1"].font = Font(italic=True, color="64748B", size=9)

    return wb, len(rows)


def generate(query_df, with_pivots=True):
    """Devuelve los bytes del .xlsx. Con with_pivots inyecta las PivotTables nativas."""
    df = query_df(FLAT_SQL)
    if df is None:
        raise RuntimeError("No hay conexión a la base de datos para exportar.")
    wb, nrows = build_base_workbook(df, with_pivots=with_pivots)
    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    if with_pivots:
        try:
            from app import pivots
        except ImportError:
            import pivots
        data = pivots.inject(data, df, DATA_COLUMNS, PIVOTS)
    return data, nrows
